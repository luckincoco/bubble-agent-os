#!/usr/bin/env python3
"""
华瑞隆进销存 Excel → Bubble Agent OS 批量导入脚本
读取 Excel 各工作表，转换为结构化 Bubble 记忆，通过 /api/import 批量写入。
"""
import json
import urllib.request
import openpyxl
from datetime import datetime

# ── 配置 (从环境变量读取) ──────────────────────────────────────────────
import os
EXCEL_PATH = os.environ.get("EXCEL_PATH", "")
API_BASE = os.environ.get("API_BASE", "")
API_KEY = os.environ.get("API_KEY", "")
SPACE_ID = os.environ.get("SPACE_ID", "")
BATCH_SIZE = 20  # 每批导入条数

if not all([EXCEL_PATH, API_BASE, API_KEY, SPACE_ID]):
    print("Required env vars: EXCEL_PATH, API_BASE, API_KEY, SPACE_ID")
    exit(1)


def fmt_date(v):
    """日期格式化"""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return str(v) if v else ""


def fmt_month(v):
    """月份格式化"""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m")
    return str(v) if v else ""


def fmt_num(v, decimals=2):
    """数值格式化"""
    if v is None or v == "":
        return 0
    try:
        return round(float(v), decimals)
    except (ValueError, TypeError):
        return 0


def send_batch(bubbles):
    """发送一批 bubble 到 API"""
    payload = json.dumps({
        "bubbles": bubbles,
        "spaceId": SPACE_ID,
    }, ensure_ascii=False).encode("utf-8")

    req = urllib.request.Request(
        f"{API_BASE}/api/import",
        data=payload,
        headers={
            "Content-Type": "application/json",
            "X-API-Key": API_KEY,
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = json.loads(resp.read())
            return result.get("created", 0)
    except Exception as e:
        print(f"  [ERROR] {e}")
        return 0


# ── 解析各工作表 ──────────────────────────────────────────────────────

def parse_purchases(ws):
    """解析采购录入表"""
    bubbles = []
    last_date = None
    last_order = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        # 列: A日期 B单号 C供应商 D商品代码 E品牌 F商品名称 G规格 H件重 I件数 J吨位 K单价 L金额 M发票 N付款 O项目 P辅助月
        date_val = row[0] or last_date
        order_no = row[1] or last_order
        supplier = row[2] or ""
        product_code = row[3] or ""
        brand = row[4] or ""
        product_name = row[5] or ""
        spec = row[6] or ""
        bundle_weight = fmt_num(row[7], 3)
        quantity = fmt_num(row[8], 0)
        tonnage = fmt_num(row[9], 3)
        unit_price = fmt_num(row[10])
        amount = fmt_num(row[11])
        invoice_status = row[12] or ""
        payment_status = row[13] or ""
        project = row[14] or ""

        if date_val:
            last_date = date_val
        if order_no:
            last_order = order_no

        # 跳过无效行
        if not tonnage and not amount:
            continue
        if not product_name and not product_code:
            continue

        date_str = fmt_date(date_val)
        month_str = fmt_month(date_val)

        title = f"{date_str} 采购 {supplier} {brand}{product_name} {spec} {tonnage}吨 ¥{amount:,.2f}"
        content = (
            f"{date_str}，从供应商{supplier}采购{brand}品牌{product_name}，"
            f"规格{spec}，件重{bundle_weight}吨/件×{int(quantity)}件={tonnage}吨，"
            f"单价{unit_price}元/吨，金额{amount:,.2f}元。"
            f"入库单号{order_no}。"
        )
        if project:
            content += f"关联项目：{project}。"
        if payment_status:
            content += f"付款状态：{payment_status}。"

        tags = ["采购", "进销存"]
        if supplier:
            tags.append(supplier)
        if brand:
            tags.append(brand)
        if product_name:
            tags.append(product_name)
        if project:
            tags.append(project)
        if month_str:
            tags.append(month_str)

        bubbles.append({
            "ref": f"purchase_{order_no}_{spec}",
            "type": "event",
            "title": title,
            "content": content,
            "metadata": {
                "category": "采购",
                "date": date_str,
                "month": month_str,
                "orderNo": str(order_no),
                "supplier": supplier,
                "brand": brand,
                "productName": product_name,
                "productCode": str(product_code),
                "spec": str(spec),
                "bundleWeight": bundle_weight,
                "quantity": int(quantity),
                "tonnage": tonnage,
                "unitPrice": unit_price,
                "amount": amount,
                "invoiceStatus": invoice_status,
                "paymentStatus": payment_status,
                "project": project,
            },
            "tags": tags,
            "source": "excel",
            "confidence": 1.0,
            "pinned": False,
        })
    return bubbles


def parse_sales(ws):
    """解析销售录入表"""
    bubbles = []
    last_date = None
    last_order = None
    last_supplier = None
    last_customer = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        # 列: A日期 B单号 C供应商 D客户/项目 E商品代码 F品牌 G商品名称 H规格 I件数 J吨位 K销售单价 L销售金额 M成本价(自动) N成本价(手动) O采购成本 P单笔毛利 Q款项状态 R物流商 S辅助月
        date_val = row[0] or last_date
        order_no = row[1] or last_order
        supplier = row[2] or last_supplier
        customer = row[3] or last_customer
        product_code = row[4] or ""
        brand = row[5] or ""
        product_name = row[6] or ""
        spec = row[7] or ""
        quantity = fmt_num(row[8], 0)
        tonnage = fmt_num(row[9], 3)
        sale_price = fmt_num(row[10])
        sale_amount = fmt_num(row[11])
        cost_auto = fmt_num(row[12])
        cost_manual = fmt_num(row[13])
        cost_amount = fmt_num(row[14])
        profit = fmt_num(row[15])
        payment_status = row[16] or ""
        logistics = row[17] or ""

        if date_val:
            last_date = date_val
        if order_no:
            last_order = order_no
        if supplier:
            last_supplier = supplier
        if customer:
            last_customer = customer

        if not tonnage and not sale_amount:
            continue
        if not product_name and not product_code:
            continue

        date_str = fmt_date(date_val)
        month_str = fmt_month(date_val)
        cost_price = cost_manual if cost_manual else cost_auto

        title = f"{date_str} 销售 {customer} {brand}{product_name} {spec} {tonnage}吨 ¥{sale_amount:,.2f}"
        content = (
            f"{date_str}，向{customer}销售{brand}品牌{product_name}，"
            f"规格{spec}，{int(quantity)}件={tonnage}吨，"
            f"销售单价{sale_price}元/吨，销售金额{sale_amount:,.2f}元。"
            f"成本价{cost_price}元/吨，采购成本{cost_amount:,.2f}元，"
            f"毛利{profit:,.2f}元。"
            f"销售单号{order_no}，供货商{supplier}。"
        )
        if logistics:
            content += f"物流商：{logistics}。"

        tags = ["销售", "进销存"]
        if customer:
            tags.append(customer)
        if supplier:
            tags.append(supplier)
        if brand:
            tags.append(brand)
        if product_name:
            tags.append(product_name)
        if month_str:
            tags.append(month_str)

        bubbles.append({
            "ref": f"sale_{order_no}_{spec}",
            "type": "event",
            "title": title,
            "content": content,
            "metadata": {
                "category": "销售",
                "date": date_str,
                "month": month_str,
                "orderNo": str(order_no),
                "supplier": supplier,
                "customer": customer,
                "brand": brand,
                "productName": product_name,
                "productCode": str(product_code),
                "spec": str(spec),
                "quantity": int(quantity),
                "tonnage": tonnage,
                "salePrice": sale_price,
                "saleAmount": sale_amount,
                "costPrice": cost_price,
                "costAmount": cost_amount,
                "profit": profit,
                "paymentStatus": payment_status,
                "logistics": logistics,
                "project": customer,
            },
            "tags": tags,
            "source": "excel",
            "confidence": 1.0,
            "pinned": False,
        })
    return bubbles


def parse_logistics(ws):
    """解析物流录入表"""
    bubbles = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # 列: A装车日期 B运单号 C托运公司 D目的地/项目 E车牌号 F司机 G司机电话 H吨位 I运费 J吊费 K费用合计 L结算状态 M辅助月
        date_val = row[0]
        order_no = row[1] or ""
        carrier = row[2] or ""
        destination = row[3] or ""
        plate = row[4] or ""
        driver = row[5] or ""
        driver_phone = row[6] or ""
        tonnage = fmt_num(row[7], 1)
        freight = fmt_num(row[8])
        lifting = fmt_num(row[9])
        total_fee = fmt_num(row[10])
        settlement = row[11] or ""

        if not date_val or (not tonnage and not total_fee):
            continue

        date_str = fmt_date(date_val)
        month_str = fmt_month(date_val)

        title = f"{date_str} 物流 {carrier} → {destination} {tonnage}吨 ¥{total_fee:,.0f}"
        content = (
            f"{date_str}，托运公司{carrier}，目的地{destination}，"
            f"司机{driver}，{tonnage}吨，"
            f"运费{freight}元+吊费{lifting}元=合计{total_fee}元。"
            f"运单号{order_no}。"
        )

        tags = ["物流", "进销存"]
        if carrier:
            tags.append(carrier)
        if destination:
            tags.append(destination)
        if month_str:
            tags.append(month_str)

        bubbles.append({
            "ref": f"logistics_{order_no}_{date_str}",
            "type": "event",
            "title": title,
            "content": content,
            "metadata": {
                "category": "物流",
                "date": date_str,
                "month": month_str,
                "orderNo": str(order_no),
                "carrier": carrier,
                "destination": destination,
                "plateNumber": str(plate),
                "driver": driver,
                "driverPhone": str(driver_phone),
                "tonnage": tonnage,
                "freight": freight,
                "liftingFee": lifting,
                "totalFee": total_fee,
                "settlement": settlement,
            },
            "tags": tags,
            "source": "excel",
            "confidence": 1.0,
            "pinned": False,
        })
    return bubbles


def parse_payments(ws):
    """解析收付款记录表"""
    bubbles = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # 列: A日期 B单据号 C类型 D对象 E关联项目 F金额 G方式 H摘要
        date_val = row[0]
        doc_no = row[1] or ""
        pay_type = row[2] or ""
        target = row[3] or ""
        project = row[4] or ""
        amount = fmt_num(row[5])
        method = row[6] or ""
        remark = row[7] or ""

        if not date_val or not amount:
            continue

        date_str = fmt_date(date_val)
        month_str = fmt_month(date_val)

        type_label = "付款" if pay_type == "付款" else "收款"
        title = f"{date_str} {type_label} {target} ¥{amount:,.2f}"
        content = (
            f"{date_str}，{type_label}给{target}，金额{amount:,.2f}元，"
            f"方式：{method}。"
        )
        if project:
            content += f"关联项目：{project}。"
        if remark:
            content += f"备注：{remark}。"

        tags = [type_label, "进销存", "收付款"]
        if target:
            tags.append(target)
        if project:
            tags.append(project)
        if month_str:
            tags.append(month_str)

        bubbles.append({
            "ref": f"payment_{date_str}_{target}_{amount}",
            "type": "event",
            "title": title,
            "content": content,
            "metadata": {
                "category": type_label,
                "date": date_str,
                "month": month_str,
                "docNo": str(doc_no),
                "type": pay_type,
                "target": target,
                "project": project,
                "amount": amount,
                "method": method,
                "remark": remark,
            },
            "tags": tags,
            "source": "excel",
            "confidence": 1.0,
            "pinned": False,
        })
    return bubbles


def parse_entities(wb):
    """解析基础信息表（供应商、客户项目、产品信息）生成实体 bubble"""
    bubbles = []

    # 供应商信息
    ws = wb["供应商信息"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[1]
        if not name:
            continue
        brands = row[2] or ""
        address = row[3] or ""
        contact = row[4] or ""
        phone = row[5] or ""
        bubbles.append({
            "ref": f"supplier_{name}",
            "type": "entity",
            "title": f"供应商：{name}",
            "content": (
                f"供应商{name}，经销品牌：{brands}，提货地址：{address}，"
                f"联系人：{contact}，电话：{phone}。"
            ),
            "metadata": {
                "entityType": "供应商",
                "name": name,
                "brands": brands,
                "address": address,
                "contact": contact,
                "phone": str(phone),
            },
            "tags": ["供应商", "基础信息", name],
            "source": "excel",
            "confidence": 1.0,
            "pinned": True,
        })

    # 客户与项目
    ws = wb["客户与项目"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[1]
        if not name:
            continue
        contract = row[2] or ""
        address = row[3] or ""
        builder = row[4] or ""
        developer = row[5] or ""
        contact = row[6] or ""
        phone = row[7] or ""
        status = row[8] or ""
        bubbles.append({
            "ref": f"project_{name}",
            "type": "entity",
            "title": f"项目：{name}",
            "content": (
                f"工程项目{name}，合同编号{contract}，工程地址：{address}，"
                f"施工单位：{builder}，建设单位：{developer}，"
                f"联系人：{contact}（{phone}），状态：{status}。"
            ),
            "metadata": {
                "entityType": "项目",
                "name": name,
                "contract": contract,
                "address": address,
                "builder": builder,
                "developer": developer,
                "contact": contact,
                "phone": str(phone),
                "status": status,
            },
            "tags": ["项目", "客户", "基础信息", name],
            "source": "excel",
            "confidence": 1.0,
            "pinned": True,
        })

    return bubbles


def parse_dashboard(wb):
    """解析经营仪表盘和年度汇总，生成综合记忆"""
    bubbles = []

    # 经营仪表盘
    ws = wb["经营仪表盘"]
    row4 = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    sales_total = fmt_num(row4[0])
    purchase_total = fmt_num(row4[1])
    margin_rate = fmt_num(row4[2], 4)
    receivable = fmt_num(row4[3])
    payable = fmt_num(row4[4])

    bubbles.append({
        "ref": "dashboard_2026",
        "type": "synthesis",
        "title": "2026年华瑞隆经营概览",
        "content": (
            f"2026年度华瑞隆经营数据：年度销售额{sales_total:,.2f}元，"
            f"年度采购额{purchase_total:,.2f}元，"
            f"综合毛利率{margin_rate*100:.2f}%，"
            f"应收账款余额{receivable:,.2f}元，"
            f"应付账款余额{payable:,.2f}元。"
        ),
        "metadata": {
            "category": "经营概览",
            "year": 2026,
            "salesTotalAmount": sales_total,
            "purchaseTotalAmount": purchase_total,
            "grossMarginRate": margin_rate,
            "receivableBalance": receivable,
            "payableBalance": payable,
        },
        "tags": ["经营概览", "2026", "仪表盘", "进销存"],
        "source": "excel",
        "confidence": 1.0,
        "pinned": True,
    })

    # 年度汇总（月度数据）
    ws = wb["年度汇总"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        month = row[0]
        if not month or not str(month).startswith("2026"):
            continue
        p_tons = fmt_num(row[1], 3)
        p_amount = fmt_num(row[2])
        s_tons = fmt_num(row[3], 3)
        s_amount = fmt_num(row[4])
        gross_profit = fmt_num(row[5])
        margin = fmt_num(row[6], 4)

        if not p_amount and not s_amount:
            continue

        bubbles.append({
            "ref": f"monthly_{month}",
            "type": "synthesis",
            "title": f"{month} 月度汇总",
            "content": (
                f"{month}月度汇总：采购{p_tons}吨/{p_amount:,.2f}元，"
                f"销售{s_tons}吨/{s_amount:,.2f}元，"
                f"毛利{gross_profit:,.2f}元，毛利率{margin*100:.2f}%。"
            ),
            "metadata": {
                "category": "月度汇总",
                "month": str(month),
                "purchaseTons": p_tons,
                "purchaseAmount": p_amount,
                "salesTons": s_tons,
                "salesAmount": s_amount,
                "grossProfit": gross_profit,
                "grossMarginRate": margin,
            },
            "tags": ["月度汇总", str(month), "进销存"],
            "source": "excel",
            "confidence": 1.0,
            "pinned": True,
        })

    # 应收账款
    ws = wb["应收账款"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if not name or name == 0:
            continue
        total_sales = fmt_num(row[1])
        logistics = fmt_num(row[2])
        receivable_total = fmt_num(row[3])
        received = fmt_num(row[4])
        outstanding = fmt_num(row[5])

        if not total_sales:
            continue

        bubbles.append({
            "ref": f"receivable_{name}",
            "type": "synthesis",
            "title": f"应收账款：{name}",
            "content": (
                f"项目{name}应收账款：累计销售额{total_sales:,.2f}元，"
                f"物流费合计{logistics:,.2f}元，"
                f"应收总额{receivable_total:,.2f}元，"
                f"已回款{received:,.2f}元，"
                f"未回款余额{outstanding:,.2f}元。"
            ),
            "metadata": {
                "category": "应收账款",
                "project": name,
                "totalSales": total_sales,
                "logisticsFee": logistics,
                "receivableTotal": receivable_total,
                "received": received,
                "outstanding": outstanding,
            },
            "tags": ["应收账款", "对账", name, "进销存"],
            "source": "excel",
            "confidence": 1.0,
            "pinned": True,
        })

    # 应付账款
    ws = wb["应付账款"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if not name or name == 0:
            continue
        total_purchase = fmt_num(row[1])
        paid = fmt_num(row[2])
        unpaid = fmt_num(row[3])

        if not total_purchase and not paid:
            continue

        bubbles.append({
            "ref": f"payable_{name}",
            "type": "synthesis",
            "title": f"应付账款：{name}",
            "content": (
                f"供应商{name}应付账款：累计采购额{total_purchase:,.2f}元，"
                f"已付金额{paid:,.2f}元，"
                f"未付余额{unpaid:,.2f}元。"
            ),
            "metadata": {
                "category": "应付账款",
                "supplier": name,
                "totalPurchase": total_purchase,
                "paidAmount": paid,
                "unpaidBalance": unpaid,
            },
            "tags": ["应付账款", "对账", name, "进销存"],
            "source": "excel",
            "confidence": 1.0,
            "pinned": True,
        })

    # 利润分析
    ws = wb["利润分析"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if not name or name == 0:
            continue
        s_total = fmt_num(row[1])
        cost = row[2]
        logi = fmt_num(row[3])
        profit = row[4]
        margin = row[5]
        avg_profit = row[6]
        s_tons = fmt_num(row[7], 3)

        if not s_total:
            continue

        # Skip #VALUE! errors
        if isinstance(cost, str) and "VALUE" in cost:
            continue

        cost = fmt_num(cost)
        profit = fmt_num(profit)
        margin = fmt_num(margin, 4)
        avg_profit = fmt_num(avg_profit)

        bubbles.append({
            "ref": f"profit_{name}",
            "type": "synthesis",
            "title": f"利润分析：{name}",
            "content": (
                f"项目{name}利润分析：销售总额{s_total:,.2f}元，"
                f"采购成本{cost:,.2f}元，物流费用{logi:,.2f}元，"
                f"毛利{profit:,.2f}元，毛利率{margin*100:.2f}%，"
                f"吨均利润{avg_profit:.2f}元/吨，销售吨位{s_tons}吨。"
            ),
            "metadata": {
                "category": "利润分析",
                "project": name,
                "salesTotalAmount": s_total,
                "costAmount": cost,
                "logisticsFee": logi,
                "grossProfit": profit,
                "grossMarginRate": margin,
                "profitPerTon": avg_profit,
                "salesTons": s_tons,
            },
            "tags": ["利润分析", name, "进销存"],
            "source": "excel",
            "confidence": 1.0,
            "pinned": True,
        })

    return bubbles


# ── 主流程 ──────────────────────────────────────────────────────────────

def main():
    print(f"正在读取 Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    all_bubbles = []

    # 1. 基础实体
    entities = parse_entities(wb)
    print(f"[基础信息] {len(entities)} 条实体记录")
    all_bubbles.extend(entities)

    # 2. 经营概览/汇总/对账
    dashboard = parse_dashboard(wb)
    print(f"[经营汇总] {len(dashboard)} 条汇总记录")
    all_bubbles.extend(dashboard)

    # 3. 采购记录
    purchases = parse_purchases(wb["采购录入"])
    print(f"[采购录入] {len(purchases)} 条采购记录")
    all_bubbles.extend(purchases)

    # 4. 销售记录
    sales = parse_sales(wb["销售录入"])
    print(f"[销售录入] {len(sales)} 条销售记录")
    all_bubbles.extend(sales)

    # 5. 物流记录
    logistics = parse_logistics(wb["物流录入"])
    print(f"[物流录入] {len(logistics)} 条物流记录")
    all_bubbles.extend(logistics)

    # 6. 收付款记录
    payments = parse_payments(wb["收付款记录"])
    print(f"[收付款] {len(payments)} 条收付款记录")
    all_bubbles.extend(payments)

    print(f"\n总计 {len(all_bubbles)} 条记录，开始批量导入...")
    print(f"目标空间: {SPACE_ID} (华瑞隆)")
    print(f"批次大小: {BATCH_SIZE}\n")

    total_created = 0
    for i in range(0, len(all_bubbles), BATCH_SIZE):
        batch = all_bubbles[i:i + BATCH_SIZE]
        created = send_batch(batch)
        total_created += created
        print(f"  批次 {i // BATCH_SIZE + 1}: 发送 {len(batch)} 条, 成功 {created} 条")

    print(f"\n导入完成! 共成功导入 {total_created}/{len(all_bubbles)} 条记录到泡泡记忆系统。")


if __name__ == "__main__":
    main()
